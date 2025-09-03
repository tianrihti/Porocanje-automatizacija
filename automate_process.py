import logging
from datetime import datetime, timedelta
from typing import List, Optional, Set
import pandas as pd
from openpyxl import load_workbook
import win32com.client
import xlwings as xw
import psutil
from collections import Counter


# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

ROW_START = 6
ROW_END = 44
COLS_TO_COPY = 3
ANALYZE_ROW_START = 7
ANALYZE_ROW_END = 46
BRIZGANJE_COL_A = 1
BRIZGANJE_COL_L = 12
BRIZGANJE_COL_M = 13

class ExcelAutomation:
    def __init__(self):
        self.pregled_file = os.path.abspath('torek (1).xls')
        self.porocanje_file = os.path.abspath("poročanje proizvodnje2025.xlsm")

        self.plan_file = os.path.abspath("plan brizganja 2025 mesečni.xlsx")

        # Validate files exist
        self._validate_files()

        self.change_log_path = os.path.join(os.path.dirname(__file__), "excel_changes.log")

    def _validate_files(self):
        """Validate that all required files exist"""
        files = [self.pregled_file, self.porocanje_file, self.plan_file]
        for file in files:
            if not os.path.exists(file):
                logger.error(f"File not found: {file}")
                raise FileNotFoundError(f"File not found: {file}")

    def log_excel_change(self, file, sheet, cell, old_value, new_value, sifra=None, action=None, target_sifra=None):
        """Log changes to a text file in the script folder."""
        with open(self.change_log_path, "a", encoding="utf-8") as f:
            if action == "copy":
                f.write(
                    f"- ŠIFRO {str(sifra).upper()} SEM KOPIRAL V {str(target_sifra).upper()}\n"
                )
            elif action == "remove":
                f.write(
                    f"- ODSTRANIL SEM VREDNOST V {cell.upper()} POD {str(sifra).upper()}\n"
                )
            else:
                # fallback to old format, but without timestamp
                f.write(
                    f"- FILE: {file} | SHEET: {sheet} | CELL: {cell} | OLD: {old_value} | NEW: {new_value}\n"
                )
        
    def step1_copy_pregled_data(self) -> pd.DataFrame:
        """Step 1: Copy data from Pregled.xls"""
        logger.info("Step 1: Copying data from Pregled.xls")
        
        try:
            # Read Pregled.xls
            try:
                df = pd.read_excel(self.pregled_file, sheet_name="Sheet1", engine='xlrd')
            except ImportError:
                df = pd.read_excel(self.pregled_file, sheet_name="Sheet1", engine='openpyxl')
            
            logger.info(f"Successfully read Pregled.xls with {len(df)} rows")
            return df
            
        except Exception as e:
            logger.error(f"Error reading Pregled.xls: {e}")
            raise
    
    def step2_paste_to_porocanje(self, data_df: pd.DataFrame) -> bool:
        logger.info("Step 2: Pasting data into poročanje proizvodnje2025.xlsm (safe method)")
        try:
            wb = load_workbook(self.porocanje_file, keep_vba=True)
            ws = wb['prilepi gosoft'] if 'prilepi gosoft' in wb.sheetnames else wb.create_sheet('prilepi gosoft')
            ws.delete_rows(1, ws.max_row)
            ws.append(list(data_df.columns))
            for row in data_df.itertuples(index=False, name=None):
                ws.append(row)
            wb.save(self.porocanje_file)
            wb.close()
            logger.info(f"Successfully pasted {len(data_df)} rows to 'prilepi gosoft' sheet (safe method)")
            return True
        except Exception as e:
            logger.error(f"Error pasting data to poročanje proizvodnje2025.xlsm (safe method): {e}")
            raise

    def get_target_date(self, holidays: Optional[Set[datetime.date]] = None) -> datetime:
        """
        Get the last working day before today (or before Monday, if today is Monday).
        :param holidays: Optional set/list of dates (datetime.date) that are holidays.
        """
        if holidays is None:
            holidays = { 
                datetime(datetime.now().year, 1, 1).date(),
                datetime(datetime.now().year, 1, 2).date(),
                datetime(datetime.now().year, 2, 8).date(),
                datetime(datetime.now().year, 4, 27).date(),
                datetime(datetime.now().year, 5, 1).date(),
                datetime(datetime.now().year, 5, 2).date(), 
                datetime(datetime.now().year, 6, 25).date(),
                datetime(datetime.now().year, 8, 15).date(),
                datetime(datetime.now().year, 10, 31).date(),
                datetime(datetime.now().year, 11, 1).date(),
                datetime(datetime.now().year, 12, 25).date(),
                datetime(datetime.now().year, 12, 26).date()        
            }

        today = datetime.now().date()
        days_back = 1 if today.weekday() != 0 else 3  

        target_date = today - timedelta(days=days_back)
        while target_date.weekday() >= 5 or target_date in holidays:
            target_date -= timedelta(days=1)
        return datetime.combine(target_date, datetime.min.time())

    def get_previous_working_day(self, date: datetime.date, holidays: Optional[Set[datetime.date]] = None) -> datetime.date:
        if holidays is None:
            holidays = { 
                datetime(datetime.now().year, 1, 1).date(),
                datetime(datetime.now().year, 1, 2).date(),
                datetime(datetime.now().year, 2, 8).date(),
                datetime(datetime.now().year, 4, 27).date(),
                datetime(datetime.now().year, 5, 1).date(),
                datetime(datetime.now().year, 5, 2).date(), 
                datetime(datetime.now().year, 6, 25).date(),
                datetime(datetime.now().year, 8, 15).date(),
                datetime(datetime.now().year, 10, 31).date(),
                datetime(datetime.now().year, 11, 1).date(),
                datetime(datetime.now().year, 12, 25).date(),
                datetime(datetime.now().year, 12, 26).date()        
            }
        prev_day = date - timedelta(days=1)
        while prev_day.weekday() >= 5 or prev_day in holidays:
            prev_day -= timedelta(days=1)
        return prev_day

    def step3_find_date_in_plan(self) -> int:
        """Step 3: Find the correct date in plan sheet"""
        logger.info("Step 3: Finding date in plan sheet")
        
        target_date = self.get_target_date()
        logger.info(f"Looking for date: {target_date.strftime('%Y-%m-%d')}")
        
        try:
            wb = load_workbook(self.plan_file, data_only=True)
            ws = wb["plan"]
            
            # Check row 4 for dates
            date_found = False
            target_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=4, column=col).value
                if cell_value:
                    # Try to parse as date
                    if isinstance(cell_value, datetime):
                        cell_date = cell_value.date()
                    else:
                        try:
                            cell_date = pd.to_datetime(cell_value).date()
                        except Exception:
                            continue
                    
                    if cell_date == target_date.date():
                        target_col = col
                        date_found = True
                        logger.info(f"Found target date at column {col}")

                        # Check if the cell below contains "Fiksno"
                        fiksno_cell = ws.cell(row=5, column=col).value
                        if fiksno_cell == "Fiksno":
                            logger.info("Found 'Fiksno' below the target date")
                        else:
                            raise ValueError("Plan is not fixed yet")
                        
                        break
        
            if not date_found:
                raise ValueError(f"Target date {target_date.strftime('%Y-%m-%d')} not found in row 4")
            
            return target_col
            
        except Exception as e:
            logger.error(f"Error finding date in plan: {e}")
            raise

    def step4_copy_plan_range(self, start_col: int) -> List[List[object]]:
        """Step 4: Copy range from plan sheet"""
        logger.info(f"Step 4: Copying range from column {start_col}")

        try:
            wb = load_workbook(self.plan_file, data_only=True)
            ws = wb["plan"]

            copied_data = [
                [ws.cell(row=row, column=start_col + col_offset).value for col_offset in range(COLS_TO_COPY)]
                for row in range(ROW_START, ROW_END + 1)
            ]

            logger.info(f"Copied range from column {start_col} to {start_col+2}, rows {ROW_START} to {ROW_END}")
            return copied_data

        except Exception as e:
            logger.error(f"Error copying plan range: {e}")
            raise
    
    def step5_paste_to_brizganje(self, copied_data: List[List[object]]):
        logger.info("Step 5: Pasting data into 'brizganje izračun' sheet")
        try:
            wb = load_workbook(self.porocanje_file, keep_vba=True)
            ws = wb["brizganje izračun"]

            # Get the target date (last working day before today)
            target_date = self.get_target_date().date()
            # Now get the previous working day before that
            prev_working_day = self.get_previous_working_day(target_date)

            logger.info(f"Looking for date: {prev_working_day.strftime('%Y-%m-%d')}")

            # Search row 4, starting from column D (index 4)
            target_col = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=4, column=col).value
                if cell_value:
                    try:
                        cell_date = cell_value.date() if isinstance(cell_value, datetime) else pd.to_datetime(cell_value).date()
                    except Exception:
                        continue

                    if cell_date == prev_working_day:
                        target_col = col
                        logger.info(f"Found target date in column {col}")
                        break

            if target_col is None:
                raise ValueError(f"Target date {prev_working_day.strftime('%Y-%m-%d')} not found in row 4")

            paste_col = target_col - 1  # One column to the left

            logger.info(f"Pasting values into column {paste_col}")

            # Clear the target range
            clear_range = ws[f"{ws.cell(row=4, column=paste_col).column_letter}4:{ws.cell(row=4+len(copied_data), column=paste_col+2).column_letter}{4+len(copied_data)}"]
            for row in clear_range:
                for cell in row:
                    cell.value = None

            # Paste copied_data into brizganje izracun sheet
            start_row = 4
            for i, row_data in enumerate(copied_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=start_row + i, column=paste_col + j, value=value)
            wb.save(self.porocanje_file)
            wb.close()
            logger.info("Successfully pasted values to 'brizganje izračun' sheet")
        except Exception as e:
            logger.error(f"Error in Step 5: {e}")
            raise

    def step6_analyze_brizganje(self) -> List[str]:
        try:
            logger.info("Step 6: Analyzing rows 7 to 46 in 'brizganje izračun'")

            # Load workbook with formula results
            wb = load_workbook(self.porocanje_file, data_only=True)
            ws = wb["brizganje izračun"]

            saved_texts = []
            
            logger.setLevel(logging.DEBUG)
            for row in range(ANALYZE_ROW_START, ANALYZE_ROW_END + 1):
                cell_a = ws.cell(row=row, column=BRIZGANJE_COL_A).value  # Column A
                if not cell_a:
                    continue  # Skip if column A is empty

                cell_l = ws.cell(row=row, column=BRIZGANJE_COL_L).value  # Column L (12)
                if cell_l in (None, ""):
                    continue  # Skip if column L is empty or zero

                cell_m = ws.cell(row=row, column=BRIZGANJE_COL_M).value  # Column M (13)
                # Try converting to float if possible
                try:
                    if isinstance(cell_m, str):
                        # Remove euro sign and replace comma with dot, if needed
                        cleaned = cell_m.replace("€", "").replace(",", ".").strip()
                        value_m = float(cleaned)
                    else:
                        value_m = float(cell_m)
                except (TypeError, ValueError):
                    continue  # Skip if not a number

                if value_m > 25:
                    saved_texts.append(str(cell_a))  # Save value from column A

            logger.info(f"Saved texts: {saved_texts}")
            return saved_texts

        except Exception as e:
            logger.error(f"Error in Step 6: {e}")
            raise

    def recalc_excel(self):
        max_retries = 3
        retry_delay = 5
        for attempt in range(max_retries):
            try:
                logger.info(f"Recalculating Excel (Attempt {attempt + 1})")
                self.kill_excel_processes()  # Only kill here, not in every step
                app = xw.App(visible=False)
                wb = app.books.open(self.porocanje_file)
                wb.app.calculate()
                wb.save()
                wb.close()
                app.quit()
                logger.info("Excel recalculation completed successfully")
                return
            except Exception as e:
                logger.error(f"Error in recalc_excel (Attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    logger.info(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                else:
                    logger.error("Max retries reached. Unable to recalculate Excel.")
                    raise
            finally:
                try:
                    if 'app' in locals():
                        app.quit()
                except Exception:
                    pass
                self.kill_excel_processes()
        raise Exception("Failed to recalculate Excel after multiple attempts")

    def kill_excel_processes(self):
        logger.info("Attempting to kill all Excel processes")
        for proc in psutil.process_iter(['name']):
            try:
                if proc.info['name'] and proc.info['name'].lower() in ['excel.exe', 'xlview.exe']:
                    proc.kill()
                    logger.info(f"Killed Excel process: {proc.pid}")
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        time.sleep(2)  # Wait for 2 seconds after killing processes

    def step7_process_saved_texts(self, saved_texts):
        logger.info("Step 7: Processing saved texts")
        excel = None
        try:
            self.kill_excel_processes()  # Only kill before win32com usage
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(self.porocanje_file)
            izbor_sheet = wb.Worksheets("izbor")
            list2_sheet = wb.Worksheets("List2")
            brizganje_izracun_sheet = wb.Worksheets("brizganje izračun")  
            
            # Delete all existing shapes (images) in the range A7:M44
            for shape in brizganje_izracun_sheet.Shapes:
                if (7 <= shape.TopLeftCell.Row <= 44 and 
                    1 <= shape.TopLeftCell.Column <= 13): 
                    shape.Delete()

            for text in saved_texts:
                logger.info(f"Processing text: {text}")

                # a. Filter data in Izbor sheet manually
                last_row = izbor_sheet.Cells(izbor_sheet.Rows.Count, "F").End(-4162).Row
                filtered_data = []

                for row in range(1, last_row + 1):
                    cell_value = izbor_sheet.Cells(row, 27).Value  # Column AA
                    if cell_value == text or row == 1:  # Include header row
                        row_data = [izbor_sheet.Cells(row, col).Value for col in range(6, 14)]  # Columns F to M
                        filtered_data.append(row_data)

                # b. Clear List2 target range
                last_row_list2 = list2_sheet.Cells(list2_sheet.Rows.Count, "T").End(-4162).Row
                list2_sheet.Range(f"T2:AA{last_row_list2}").ClearContents()

                # c. Paste filtered data into List2 sheet
                for i, row_data in enumerate(filtered_data):
                    for j, value in enumerate(row_data):
                        list2_sheet.Cells(i + 1, 20 + j).Value = value  # Start from column T (20th column)

                # d. Execute macro (gumb1)
                excel.Run("sortiraj")

                # Step 8: Copy processed data
                self.step8_copy_processed_data(list2_sheet)

                # Step 9: Paste as image in "brizganje izracun" sheet
                self.step9_paste_as_image(brizganje_izracun_sheet, text)

                excel.CutCopyMode = False  # Clear clipboard

            # Set the height of rows 7 to 44 to 16.5 if they don't contain an image
            for row in range(7, 45):
                has_image = False
                for shape in brizganje_izracun_sheet.Shapes:
                    if shape.TopLeftCell.Row == row:
                        has_image = True
                        break
                if not has_image:
                    brizganje_izracun_sheet.Rows(row).RowHeight = 16.5
            logger.info("Adjusted heights of rows without images to 16.5")

            wb.Save()
            wb.Close()
            excel.Quit()
            logger.info("Step 7,8 and 9 completed successfully")

        except Exception as e:
            logger.error(f"Error in Step 7: {e}")
            raise
        finally:
            if excel:
                try:
                    excel.Quit()
                except Exception:
                    pass
            self.kill_excel_processes()
    
    def enable_macros(self, wb):
        try:
            # Check if there's a security alert
            if wb.ReadOnly:
                # Try to enable content
                app = wb.Application
                app.DisplayAlerts = False
                app.EnableEvents = False
                app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
                app.Run("Auto_Open")  # This usually enables content
                app.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception as e:
            logger.error(f"Failed to enable macros: {e}")
    
    def step8_copy_processed_data(self, list2_sheet):
        """Step 8: Copy processed data from List2 sheet"""
        logger.info("Step 8: Copying processed data from List2")

        # Find the last row with data in column C
        last_row = 8  # Default to row 8
        for row in range(9, 28):  # Check rows 9 and 10
            if list2_sheet.Cells(row, 3).Value:  # Column C
                last_row = row

        # Construct the range to copy
        range_to_copy = list2_sheet.Range(f"B1:L{last_row}")
        range_to_copy.CopyPicture(Appearance=1, Format=2)  # Copy as picture

        logger.info(f"Successfully copied range B1:L{last_row} from List2 as picture")

    def step9_paste_as_image(self, brizganje_izracun_sheet, text):
        """Step 9: Paste as image in 'brizganje izračun' sheet"""
        logger.info(f"Step 9: Pasting as image for text '{text}'")

        last_row_brizganje = brizganje_izracun_sheet.Cells(brizganje_izracun_sheet.Rows.Count, "A").End(-4162).Row
        target_row = None
        for row in range(1, last_row_brizganje + 1):
            if brizganje_izracun_sheet.Cells(row, 1).Value == text:
                target_row = row + 1  # Go one row below
                break

        if target_row:
            target_cell = brizganje_izracun_sheet.Cells(target_row, 1)
            brizganje_izracun_sheet.Paste(target_cell, Link=False)
    
            # Get the last pasted shape (which should be our image)
            last_shape = brizganje_izracun_sheet.Shapes(brizganje_izracun_sheet.Shapes.Count)
    
            # Adjust row height to fit the image
            image_height = last_shape.Height
            brizganje_izracun_sheet.Rows(target_row).RowHeight = image_height + 5
    
            logger.info(f"Successfully pasted image for text '{text}' at row {target_row} and adjusted row height")
            # Add this line to check if the shape is actually there
            logger.info(f"Shape count after pasting: {brizganje_izracun_sheet.Shapes.Count}")
        else:
            logger.warning(f"Could not find row for text '{text}' in 'brizganje izračun' sheet")

    def scan_brizganje_errors(self):
        logger.info("Scanning for errors in 'brizganje izračun' column H, rows 7–45")
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(self.porocanje_file)
        ws = wb.Worksheets("brizganje izračun")
        izbor_sheet = wb.Worksheets("izbor")
        excel_error_prefixes = ("#",)
        processed_keys = set()
        fixed_errors = []
        try:
            last_row_izbor = izbor_sheet.Cells(izbor_sheet.Rows.Count, "F").End(-4162).Row
            for row in range(7, 46):
                cell_h = ws.Cells(row, 8).Value  # Column H

                # Check for Excel error values (including COM error constants)
                if cell_h is None:
                    continue
                if isinstance(cell_h, str):
                    if not cell_h.startswith(excel_error_prefixes):
                        continue
                else:
                    # For non-string error values (like Excel error constants)
                    try:
                        # Excel error codes are usually negative integers
                        if isinstance(cell_h, (int, float)) and cell_h < 0:
                            pass  # treat as error
                        else:
                            continue
                    except Exception:
                        continue

                cell_i = ws.Cells(row, 9).Value  # Column I
                if not cell_i:
                    old_value = ws.Cells(row, 5).Value
                    ws.Cells(row, 5).Value = None  # Delete value in column E
                    sifra = ws.Cells(row, 1).Value  # Column A
                    self.log_excel_change(
                        self.porocanje_file, "brizganje izračun", f"E{row}", old_value, None, sifra=sifra, action="remove"
                    )
                    continue

                text_key = ws.Cells(row, 1).Value  # Column A
                if not text_key or text_key in processed_keys:
                    continue

                # Find the most common value in column S for this text_key in Izbor
                s_values = []
                for izbor_row in range(2, last_row_izbor + 1):  # skip header row
                    cell_aa = izbor_sheet.Cells(izbor_row, 27).Value  # Column AA
                    if cell_aa == text_key:
                        value_s = izbor_sheet.Cells(izbor_row, 19).Value  # Column S
                        if value_s not in (None, ""):
                            s_values.append(value_s)
                if s_values:
                    freq = Counter(s_values)
                    best_value, _ = freq.most_common(1)[0]
                    logger.info(f"For text_key '{text_key}': best_value in column S: {best_value}")
                else:
                    best_value = None
                    logger.info(f"For text_key '{text_key}': no values found in column S")
                    self.log_warning(f"Za {text_key}: nisem nasel vredonst {best_value} v stolpcu S na izboru")
                    processed_keys.add(text_key)
                    continue

                # Find the first row in brizganje izračun where column C matches best_value
                values_to_copy = []
                found_row = None
                for c_row in range(7, 46):
                    cell_c = ws.Cells(c_row, 3).Value  # Column C
                    try:
                        # Try to compare as numbers if possible
                        if cell_c is not None and best_value is not None:
                            try:
                                if float(cell_c) == float(best_value):
                                    found_row = c_row
                            except (ValueError, TypeError):
                                # Fallback to string comparison
                                if str(cell_c).strip() == str(best_value).strip():
                                    found_row = c_row
                    except Exception:
                        continue

                    if found_row:
                        # Copy C,D,E from c_row and c_row+1
                        for r in [c_row, c_row + 1]:
                            row_values = [ws.Cells(r, col).Value for col in range(3, 6)]
                            values_to_copy.extend(row_values)
                        break

                if not values_to_copy or found_row is None:
                    logger.warning(f"No matching row found for best_value '{best_value}' for text_key '{text_key}'")
                    self.log_warning(f"Ni sifre: '{best_value}' od: '{text_key}' v stolpcu C na brizganju izracuna")
                    processed_keys.add(text_key)
                    continue

                idx = 0
                for r in [row, row + 1]:
                    for col in range(3, 6):
                        cell_ref = f"{chr(64+col)}{r}"
                        old_value = ws.Cells(r, col).Value
                        new_value = values_to_copy[idx]
                        ws.Cells(r, col).Value = new_value
                        idx += 1

                source_sifra = ws.Cells(found_row, 1).Value
                target_sifra = ws.Cells(row, 1).Value
                self.log_excel_change(
                    self.porocanje_file, "brizganje izračun", f"C{row}", None, None, sifra=source_sifra, action="copy", target_sifra=target_sifra
                )

                processed_keys.add(text_key)
                fixed_errors.append({'row': row, 'text_key': text_key, 'best_value': best_value})

            logger.info(f"scan_brizganje_errors completed. Fixed {len(fixed_errors)} errors.")
        except Exception as e:
            logger.error(f"Error in scan_brizganje_errors: {e}")
            raise
        finally:
            wb.Close()
            excel.Quit()
    
    def log_warning(self, message):
        with open(self.change_log_path, "a", encoding="utf-8") as f:
            f.write(f"WARNING - {message}\n")
                    
if __name__ == "__main__":
    # Clear the log file before running
    log_path = os.path.join(os.path.dirname(__file__), "excel_changes.log")
    with open(log_path, "w", encoding="utf-8") as f:
        pass  # This clears the file

    # Create automation instance
    automation = ExcelAutomation()
    try:
        automation.kill_excel_processes()
        # Run step 1
        pregled_data = automation.step1_copy_pregled_data()

        # Run step 2
        automation.step2_paste_to_porocanje(pregled_data)

        # Run step 3
        target_col = automation.step3_find_date_in_plan()

        # Run step 4
        plan_range_data = automation.step4_copy_plan_range(target_col)

        # Run step 5
        automation.step5_paste_to_brizganje(plan_range_data)

        automation.recalc_excel()

        automation.scan_brizganje_errors()

        # Run step 6
        saved_texts = automation.step6_analyze_brizganje()

        # Run step 7
        automation.step7_process_saved_texts(saved_texts)
        
    except Exception as e:
        logger.error((f"An error occurred: {e}"))
    finally:
        automation.kill_excel_processes()
        logger.info("Script execution finished")
